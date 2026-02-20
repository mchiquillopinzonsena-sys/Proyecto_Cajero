import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import hashlib
import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import pandas as pd
import os
from tkinter import messagebox, filedialog
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table as PDFTable, TableStyle, Image
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle



conexion = sqlite3.connect("finanzas.db")
cursor = conexion.cursor()

try:
    cursor.execute("ALTER TABLE usuarios ADD COLUMN nombre TEXT DEFAULT ''")
    cursor.execute("ALTER TABLE usuarios ADD COLUMN apellido TEXT DEFAULT ''")
    conexion.commit()
except:
    pass

def crear_tablas():
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        documento TEXT UNIQUE NOT NULL,
        contraseña TEXT NOT NULL,
        nombre TEXT,
        apellido TEXT
    )
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS movimientos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        usuario_id INTEGER,
        tipo TEXT NOT NULL,
        descripcion TEXT NOT NULL,
        monto REAL NOT NULL,
        fecha TEXT NOT NULL,
        FOREIGN KEY (usuario_id) REFERENCES usuarios(id)
    )
    """)
    conexion.commit()

crear_tablas()


def encriptar(password):
    return hashlib.sha256(password.encode()).hexdigest()

def verificar_login(documento, contraseña):
    cursor.execute(
        "SELECT id, nombre, apellido FROM usuarios WHERE documento=? AND contraseña=?",
        (documento, encriptar(contraseña))
    )
    return cursor.fetchone()

def recuperar_contrasena():
    ventana_rec = tk.Toplevel()
    ventana_rec.title("Recuperar / Cambiar contraseña")
    ventana_rec.geometry("350x250")
    ventana_rec.configure(bg="#0F172A")

    tk.Label(ventana_rec, text="Documento", bg="#0F172A", fg="white").pack(pady=5)
    entry_doc_rec = tk.Entry(ventana_rec)
    entry_doc_rec.pack(pady=5)

    tk.Label(ventana_rec, text="Nueva Contraseña", bg="#0F172A", fg="white").pack(pady=5)
    entry_nueva = tk.Entry(ventana_rec, show="*")
    entry_nueva.pack(pady=5)

    tk.Label(ventana_rec, text="Confirmar Contraseña", bg="#0F172A", fg="white").pack(pady=5)
    entry_confirm = tk.Entry(ventana_rec, show="*")
    entry_confirm.pack(pady=5)

    def cambiar():
        doc = entry_doc_rec.get().strip()
        nueva = entry_nueva.get().strip()
        confirm = entry_confirm.get().strip()

        if not doc or not nueva or not confirm:
            messagebox.showerror("Error", "Complete todos los campos")
            return
        if nueva != confirm:
            messagebox.showerror("Error", "Las contraseñas no coinciden")
            return

        cursor.execute("SELECT id FROM usuarios WHERE documento=?", (doc,))
        usuario = cursor.fetchone()
        if usuario:
            cursor.execute(
                "UPDATE usuarios SET contraseña=? WHERE id=?",
                (encriptar(nueva), usuario[0])
            )
            conexion.commit()
            messagebox.showinfo("Éxito", "Contraseña cambiada correctamente")
            ventana_rec.destroy()
        else:
            messagebox.showerror("Error", "Usuario no encontrado")

    tk.Button(
        ventana_rec,
        text="Cambiar contraseña",
        bg="#16A34A",
        fg="white",
        command=cambiar
    ).pack(pady=15)


def guardar_movimiento(tipo, descripcion, monto):
    fecha = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    cursor.execute(
        "INSERT INTO movimientos (usuario_id, tipo, descripcion, monto, fecha) VALUES (?, ?, ?, ?, ?)",
        (usuario_actual, tipo, descripcion, monto, fecha)
    )
    conexion.commit()

def obtener_saldo():
    cursor.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN tipo='ingreso' THEN monto END), 0) -
            COALESCE(SUM(CASE WHEN tipo='egreso' THEN monto END), 0)
        FROM movimientos
        WHERE usuario_id=?
    """, (usuario_actual,))
    return cursor.fetchone()[0]

def obtener_historial():
    cursor.execute("""
        SELECT tipo, descripcion, monto, fecha
        FROM movimientos
        WHERE usuario_id=?
        ORDER BY id DESC
    """, (usuario_actual,))
    return cursor.fetchall()

# EXPORTAR 

def exportar_a_excel():
    try:
        cursor.execute("""
            SELECT tipo, descripcion, monto, fecha
            FROM movimientos
            WHERE usuario_id=?
        """, (usuario_actual,))
        datos = cursor.fetchall()

        if not datos:
            messagebox.showinfo("Info", "No hay movimientos para exportar")
            return

        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")]
        )

        if not archivo:
            return

        df = pd.DataFrame(datos, columns=["Tipo", "Descripción", "Monto", "Fecha"])
        df.to_excel(archivo, index=False)

        wb = load_workbook(archivo)
        ws = wb.active

        rango = f"A1:D{ws.max_row}"
        tabla = Table(displayName="TablaMovimientos", ref=rango)

        estilo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        tabla.tableStyleInfo = estilo
        ws.add_table(tabla)

        wb.save(archivo)

        os.startfile(archivo)  # Abre automáticamente el archivo

        messagebox.showinfo("Exportado", "Movimientos exportados a Excel exitosamente!")

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo exportar a Excel:\n{e}")

def exportar_a_pdf():
    try:
        cursor.execute("""
            SELECT tipo, descripcion, monto, fecha
            FROM movimientos
            WHERE usuario_id=?
        """, (usuario_actual,))
        datos = cursor.fetchall()

        if not datos:
            messagebox.showinfo("Info", "No hay movimientos para exportar")
            return

        archivo_pdf = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("Archivos PDF", "*.pdf")]
        )

        if not archivo_pdf:
            return

        doc = SimpleDocTemplate(archivo_pdf, pagesize=letter)
        elementos = []

        # Imagen superior
        img = Image("fondo.png", width=500, height=120)
        elementos.append(img)
        elementos.append(Spacer(1, 20))

        # Datos usuario
        cursor.execute("SELECT nombre, documento FROM usuarios WHERE id=?", (usuario_actual,))
        usuario = cursor.fetchone()

        estilo = ParagraphStyle(name='Normal', fontSize=12)

        elementos.append(Paragraph(f"<b>Usuario:</b> {usuario[0]}", estilo))
        elementos.append(Paragraph(f"<b>Documento:</b> {usuario[1]}", estilo))
        elementos.append(Spacer(1, 20))

        # Tabla
        datos_tabla = [["Tipo", "Descripción", "Monto", "Fecha"]]

        for tipo, desc, monto, fecha in datos:
            datos_tabla.append([tipo, desc, f"${monto:,.2f}", fecha])

        tabla = PDFTable(datos_tabla, colWidths=[80, 200, 100, 100])

        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1E293B")),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('ALIGN',(2,1),(-1,-1),'CENTER'),
            ('GRID', (0,0), (-1,-1), 1, colors.grey),
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
        ]))

        elementos.append(tabla)

        doc.build(elementos)

        os.startfile(archivo_pdf)  # Abre automáticamente

        messagebox.showinfo("Exportado", "Movimientos exportados a PDF exitosamente!")

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo exportar a PDF:\n{e}")
def iniciar_sistema():
    ventana = tk.Tk()
    ventana.title("Sistema Financiero PRO")
    ventana.geometry("900x650")
    ventana.configure(bg="#17233F")

    style = ttk.Style()
    style.theme_use("clam")
    style.configure("Treeview",
        background="#1E293B",
        foreground="white",
        rowheight=28,
        fieldbackground="#1E293B"
    )
    style.map("Treeview", background=[("selected", "#334155")])

    barra = tk.Frame(ventana, bg="#111827", height=60)
    barra.pack(fill="x")

    tk.Label(
        barra,
        text="💰 Sistema Financiero",
        bg="#111827",
        fg="white",
        font=("Segoe UI", 16, "bold")
    ).pack(side="left", padx=20)

    tk.Button(
        barra,
        text="Exportar Excel",
        bg="#22C55E",
        fg="white",
        relief="flat",
        command=exportar_a_excel
    ).pack(side="right", padx=20)

    tk.Button(
        barra,
        text="Exportar PDF",
        bg="#22C55E",
        fg="white",
        relief="flat",
        command=exportar_a_pdf
    ).pack(side="right", padx=20)

    tk.Button(
        barra,
        text="Cerrar sesión",
        bg="#DC2626",
        fg="white",
        relief="flat",
        command=lambda: [ventana.destroy(), ventana_login()]
    ).pack(side="right", padx=20)

    frame_form = tk.Frame(ventana, bg="#1E293B")
    frame_form.pack(pady=20, padx=40, fill="x")

    tk.Label(
        frame_form,
        text="Registrar Movimiento",
        bg="#1E293B",
        fg="white",
        font=("Segoe UI", 14, "bold")
    ).grid(row=0, column=0, columnspan=2, pady=10)

    tk.Label(frame_form, text="Descripción", bg="#1E293B", fg="#CBD5E1").grid(row=1, column=0, sticky="w", pady=5)
    entry_descripcion = tk.Entry(frame_form, width=40, bg="#334155", fg="white", insertbackground="white")
    entry_descripcion.grid(row=1, column=1, pady=5)

    tk.Label(frame_form, text="Monto", bg="#1E293B", fg="#CBD5E1").grid(row=2, column=0, sticky="w", pady=5)
    entry_monto = tk.Entry(frame_form, width=40, bg="#334155", fg="white", insertbackground="white")
    entry_monto.grid(row=2, column=1, pady=5)

    def agregar_ingreso():
        descripcion = entry_descripcion.get().strip()
        try:
            monto = float(entry_monto.get())
        except:
            messagebox.showerror("Error", "Monto inválido")
            return

        if not descripcion:
            messagebox.showerror("Error", "Ingrese descripción")
            return

        guardar_movimiento("ingreso", descripcion, monto)
        actualizar_saldo()
        cargar_historial()

    def agregar_egreso():
        descripcion = entry_descripcion.get().strip()
        try:
            monto = float(entry_monto.get())
        except:
            messagebox.showerror("Error", "Monto inválido")
            return

        if not descripcion:
            messagebox.showerror("Error", "Ingrese descripción")
            return

        guardar_movimiento("egreso", descripcion, monto)
        actualizar_saldo()
        cargar_historial()

    tk.Button(frame_form, text="AGREGAR INGRESO", bg="#16A34A", fg="white", command=agregar_ingreso).grid(row=3, column=0, pady=15)
    tk.Button(frame_form, text="AGREGAR EGRESO", bg="#DC2626", fg="white", command=agregar_egreso).grid(row=3, column=1, pady=15)

    frame_saldo = tk.Frame(ventana, bg="#1E293B")
    frame_saldo.pack(pady=10, padx=40, fill="x")

    label_saldo = tk.Label(frame_saldo, text="Saldo actual: $0.00", font=("Segoe UI", 18, "bold"), bg="#1E293B", fg="#FACC15")
    label_saldo.pack(pady=15)

    tabla_frame = tk.Frame(ventana, bg="#0F172A")
    tabla_frame.pack(padx=40, pady=20, fill="both", expand=True)

    columns = ("Tipo", "Descripción", "Monto", "Fecha")
    tabla = ttk.Treeview(tabla_frame, columns=columns, show="headings")

    for col in columns:
        tabla.heading(col, text=col)
        tabla.column(col, anchor="center")

    tabla.pack(fill="both", expand=True)

    def actualizar_saldo():
        saldo = obtener_saldo()
        label_saldo.config(text=f"Saldo actual: ${saldo:,.2f}")

    def cargar_historial():
        tabla.delete(*tabla.get_children())
        for tipo, desc, monto, fecha in obtener_historial():
            tabla.insert("", tk.END, values=(tipo, desc, f"${monto:,.2f}", fecha))

    actualizar_saldo()
    cargar_historial()

    ventana.mainloop()



def ventana_registro():
    reg = tk.Toplevel()
    reg.title("Registro de Usuario")
    reg.geometry("400x350")
    reg.configure(bg="#0F172A")

    tk.Label(reg, text="Registrar Nuevo Usuario", bg="#0F172A", fg="white", font=("Segoe UI", 16, "bold")).pack(pady=15)

    tk.Label(reg, text="Documento", bg="#0F172A", fg="white").pack(pady=5)
    entry_doc = tk.Entry(reg)
    entry_doc.pack(pady=5)

    tk.Label(reg, text="Nombre", bg="#0F172A", fg="white").pack(pady=5)
    entry_nombre = tk.Entry(reg)
    entry_nombre.pack(pady=5)

    tk.Label(reg, text="Apellido", bg="#0F172A", fg="white").pack(pady=5)
    entry_apellido = tk.Entry(reg)
    entry_apellido.pack(pady=5)

    tk.Label(reg, text="Contraseña", bg="#0F172A", fg="white").pack(pady=5)
    entry_pass = tk.Entry(reg, show="*")
    entry_pass.pack(pady=5)

    def guardar_registro():
        doc = entry_doc.get().strip()
        nombre = entry_nombre.get().strip()
        apellido = entry_apellido.get().strip()
        pas = entry_pass.get().strip()

        if not doc or not nombre or not apellido or not pas:
            messagebox.showerror("Error", "Complete todos los campos")
            return

        try:
            cursor.execute(
                "INSERT INTO usuarios (documento, nombre, apellido, contraseña) VALUES (?, ?, ?, ?)",
                (doc, nombre, apellido, encriptar(pas))
            )
            conexion.commit()
            messagebox.showinfo("Éxito", "Usuario registrado correctamente")
            reg.destroy()
        except:
            messagebox.showerror("Error", "Usuario ya existe o dato incorrecto")

    tk.Button(reg, text="Registrar", bg="#16A34A", fg="white", command=guardar_registro).pack(pady=15)

def ventana_login():
    login = tk.Tk()
    login.title("Login")
    login.geometry("400x420")
    login.configure(bg="#0F172A")

    tk.Label(login, text="Iniciar Sesión", bg="#0F172A", fg="white", font=("Segoe UI", 16, "bold")).pack(pady=20)

    tk.Label(login, text="Documento", bg="#0F172A", fg="white").pack()
    entry_doc = tk.Entry(login)
    entry_doc.pack(pady=5)

    tk.Label(login, text="Contraseña", bg="#0F172A", fg="white").pack()
    entry_pass = tk.Entry(login, show="*")
    entry_pass.pack(pady=5)

    def iniciar_sesion():
        doc = entry_doc.get()
        pas = entry_pass.get()
        usuario = verificar_login(doc, pas)

        if usuario:
            global usuario_actual
            usuario_actual = usuario[0]
            login.destroy()
            iniciar_sistema()
        else:
            messagebox.showerror("Error", "Datos incorrectos")

    tk.Button(login, text="Iniciar sesión", bg="#16A34A", fg="white", command=iniciar_sesion).pack(pady=10)

    tk.Button(login, text="Registrarse", bg="#2563EB", fg="white", command=ventana_registro).pack()

    tk.Button(login, text="¿Olvidaste tu contraseña?", bg="#FAD015", fg="black", command=recuperar_contrasena).pack(pady=5)

    login.mainloop()

ventana_login()
conexion.close()